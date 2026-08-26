from __future__ import annotations

import csv
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from .model import PanelItem

NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def read_panel_items(path: Path) -> list[PanelItem]:
    """Read the simple panel-list shape used by the sample Výpis prvkov workbook."""
    try:
        return _read_with_openpyxl(path)
    except ModuleNotFoundError:
        return _read_with_stdlib(path)


def write_panel_csv(items: list[PanelItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["view", "mark", "length_m", "width_m", "count", "area_m2", "note"],
        )
        writer.writeheader()
        for item in items:
            writer.writerow(item.__dict__)


def _read_with_openpyxl(path: Path) -> list[PanelItem]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    items: list[PanelItem] = []
    current_view = ""
    for row in sheet.iter_rows(values_only=True):
        first, mark, length, width, count, area, note = (list(row) + [None] * 7)[:7]
        if isinstance(first, str) and first.strip() and first.strip().lower() not in {"pohľad", "spolu", "rezerva"}:
            current_view = first.strip()
        parsed = _panel_from_cells(current_view, mark, length, width, count, area, note)
        if parsed is not None:
            items.append(parsed)
    return items


def _read_with_stdlib(path: Path) -> list[PanelItem]:
    rows = _xlsx_rows(path)
    items: list[PanelItem] = []
    current_view = ""
    for row in rows:
        first = row.get("A", "")
        if first and first.lower() not in {"pohľad", "spolu", "rezerva"}:
            current_view = first
        parsed = _panel_from_cells(
            current_view,
            row.get("B"),
            row.get("C"),
            row.get("D"),
            row.get("E"),
            row.get("F"),
            row.get("G"),
        )
        if parsed is not None:
            items.append(parsed)
    return items


def _panel_from_cells(
    view: str,
    mark: object,
    length: object,
    width: object,
    count: object,
    area: object,
    note: object,
) -> PanelItem | None:
    if not isinstance(mark, str) or not mark.strip() or mark.strip().lower() in {"označenie", "rezerva"}:
        return None
    length_m = _float_or_none(length)
    width_m = _float_or_none(width)
    count_int = _int_or_none(count)
    area_m2 = _float_or_none(area)
    if length_m is None or width_m is None or count_int is None or area_m2 is None:
        return None
    return PanelItem(
        view=view,
        mark=mark.strip(),
        length_m=length_m,
        width_m=width_m,
        count=count_int,
        area_m2=area_m2,
        note=str(note or "").strip(),
    )


def _xlsx_rows(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _shared_strings(archive)
        sheet = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    rows = []
    for row_node in sheet.findall(".//a:sheetData/a:row", NS):
        row: dict[str, str] = {}
        for cell in row_node.findall("a:c", NS):
            ref = cell.attrib.get("r", "")
            column = "".join(char for char in ref if char.isalpha())
            value_node = cell.find("a:v", NS)
            if value_node is None:
                continue
            value = value_node.text or ""
            if cell.attrib.get("t") == "s" and value:
                value = shared_strings[int(value)]
            row[column] = value.strip() if isinstance(value, str) else str(value)
        rows.append(row)
    return rows


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return [
        "".join(text.text or "" for text in item.findall(".//a:t", NS))
        for item in root.findall("a:si", NS)
    ]


def _float_or_none(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _int_or_none(value: object) -> int | None:
    number = _float_or_none(value)
    if number is None:
        return None
    return int(number)

